#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Spec Sync v1.2 – advanced sorter for aircraft specifications
(см. CHANGELOG ниже)

Author : ChatGPT o3 for Владимир, 2025-05-26
License: MIT
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import logging
import os
import re
import shutil
import sys
import threading
import time
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Callable, Optional

# ────────────────────────── 3-rd-party ───────────────────────────
import openpyxl
import yaml
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter
from PySide6.QtCore import Qt, QThread, Signal, Slot
from PySide6.QtGui import QAction, QCloseEvent, QIcon, QPixmap
from PySide6.QtWidgets import (
    QApplication,
    QFileDialog,
    QLabel,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMessageBox,
    QProgressBar,
    QStatusBar,
    QTabWidget,
    QTextEdit,
    QTreeWidget,
    QTreeWidgetItem,
    QWidget,
    QPushButton,
    QLineEdit,
    QVBoxLayout,
    QHBoxLayout,
    QRadioButton,
    QCheckBox,
)

try:                       # optional deps
    import pytesseract
    from PIL import Image
except ModuleNotFoundError:
    pytesseract = None     # type: ignore

try:
    from pdfminer.high_level import extract_text as pdf_text
except ModuleNotFoundError:
    pdf_text = None        # type: ignore

try:                       # .lnk creation
    import win32com.client
except ModuleNotFoundError:
    win32com = None        # type: ignore

# ───────────────────────── Config ────────────────────────────────
ALLOWED_EXTS = {".pdf", ".tif", ".tiff", ".jpg", ".jpeg", ".png", ".bmp", ".gif"}
INDEX_HEADERS = (
    "Abbrev", "File_Name", "Format", "Revision", "Category", "Tags",
    "Size_MB", "Modified", "SHA256", "Link",
)
REV_PAT = re.compile(
    r"(?:\b|_|\-|\()(?:(?:REV(?:ISION)?|ISSUE)[ _\-]*)"
    r"(?P<rev>[A-Z0-9]{1,3})", re.I
)
CHUNK = 1 << 20  # 1 MiB
LOG = logging.getLogger("specsync")
LOG.setLevel(logging.DEBUG)
handler = logging.StreamHandler(sys.stdout)
handler.setFormatter(logging.Formatter(
    "%(asctime)s  %(levelname)-8s  %(message)s", "%H:%M:%S"
))
LOG.addHandler(handler)


# ═════════════════════════ Engine ════════════════════════════════
class AbbrevMatcher:
    """Гибкая проверка имени файла на наличие аббревиатур."""

    def __init__(self, abbrevs: set[str], *, case_sensitive: bool):
        flags = 0 if case_sensitive else re.I
        # сортируем от длинных к коротким → минимизируем ложные срабатывания
        self.abbrev_order = sorted(abbrevs, key=len, reverse=True)
        # составляем один big-regex:  (?P<A>AMS)|(?P<B>BAC)|…
        parts = []
        self.group2abbr: dict[str, str] = {}
        for idx, abbr in enumerate(self.abbrev_order):
            safe = re.escape(abbr)
            # допускаем, что следом может идти тире/подчёркивание/цифры/скобка
            pat = rf"(?P<G{idx}>{safe})(?=[\W_0-9]|$)"
            parts.append(pat)
            self.group2abbr[f"G{idx}"] = abbr
        self.regex = re.compile("|".join(parts), flags)

    def find(self, fname: str) -> Optional[str]:
        m = self.regex.search(fname)
        if not m:
            return None
        # выясняем, какая именно группа сработала
        for g, abbr in self.group2abbr.items():
            if m.group(g):
                return abbr
        return None  # pragma: no cover


class SpecSync:
    """Главный рабочий класс – выполняется в фоновом потоке."""

    # ──────────────── init & helpers ───────────────────────────
    def __init__(
        self,
        source_roots: list[Path],
        target_root: Path,
        abbreviations_file: Path,
        rules_file: Optional[Path] = None,
        journal_file: Optional[Path] = None,
        *,
        mode: str = "copy",
        case_sensitive: bool = True,
        use_llm: bool = False,
        dry_run: bool = False,
        progress_cb: Callable[[dict], None] | None = None,
        cancel_event: threading.Event | None = None,
    ):
        self.src_roots = source_roots
        self.tgt_root = target_root
        self.mode = mode
        self.dry = dry_run
        self.use_llm = use_llm
        self.progress_cb = progress_cb or (lambda *_: None)
        self.cancel = cancel_event or threading.Event()

        self.abbrevs = self._load_abbrevs(abbreviations_file)
        self.matcher = AbbrevMatcher(self.abbrevs, case_sensitive=case_sensitive)
        self.rules = self._load_rules(rules_file)
        self.journal_sha, self.journal_names = (
            self._load_journal(journal_file) if journal_file else (set(), set())
        )

        self.rows: list[list] = []
        self.stats = dict(scanned=0, unique=0, dup=0, skipped=0, errors=0)

        self.hash_seen: dict[str, Path] = {}  # SHA-256 ⇒ first-path

    @staticmethod
    def _load_abbrevs(p: Path) -> set[str]:
        return {ln.strip() for ln in p.read_text("utf-8").splitlines() if ln.strip()}

    @staticmethod
    def _load_rules(pth: Optional[Path]) -> dict:
        if pth and pth.exists():
            return yaml.safe_load(pth.read_text("utf-8"))
        return {"rules": {}, "fallback_category": "Unknown", "case_sensitive": False}

    @staticmethod
    def _load_journal(p: Path) -> tuple[set[str], set[str]]:
        """Возвращает (set-SHA, set-filename_without_ext)."""
        sha_set, name_set = set(), set()
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        ws = wb.active
        headers = {str(c.value).strip().upper(): i for i, c in enumerate(next(ws.rows))}
        for row in ws.iter_rows(min_row=2):
            if "SHA256" in headers:
                sha_set.add(str(row[headers["SHA256"]].value))
            if "FILE_NAME" in headers:
                name_set.add(str(row[headers["FILE_NAME"]].value).upper())
        return sha_set, name_set

    # ──────────────── public run() ─────────────────────────────
    def run(self) -> None:
        try:
            self._scan()
            if not self.dry:
                self._write_index()
        except Exception:
            LOG.exception("Unexpected crash")
            self.stats["errors"] += 1
        finally:
            self.progress_cb({**self.stats, "final": True})

    # ──────────────── core scan loop ───────────────────────────
    def _scan(self) -> None:
        for root in self.src_roots:
            for path in root.rglob("*"):
                if self.cancel.is_set():
                    return
                if not path.is_file() or path.suffix.lower() not in ALLOWED_EXTS:
                    continue

                self.stats["scanned"] += 1
                try:
                    self._handle(path)
                except Exception as exc:
                    LOG.error("❌ %s : %s", path.name, exc)
                    self.stats["errors"] += 1

                if not self.stats["scanned"] % 100:
                    self.progress_cb(self.stats)

    # ──────────────── single-file processing ───────────────────
    def _handle(self, path: Path) -> None:
        abbrev = self.matcher.find(path.name)
        if not abbrev:
            self.stats["skipped"] += 1
            return

        sha = self._sha256(path)
        if sha in self.hash_seen or sha in self.journal_sha:
            LOG.debug("↩︎ dup-SHA  %s", path.name)
            self.stats["dup"] += 1
            return

        base_name = path.stem.upper()
        if (
            (self.tgt_root / abbrev / path.name).exists()
            or base_name in self.journal_names
        ):
            LOG.debug("↩︎ already indexed %s", path.name)
            self.stats["dup"] += 1
            return

        # ---------------- unique ----------------
        self.hash_seen[sha] = path
        self.stats["unique"] += 1

        dest_dir = self.tgt_root / abbrev
        dest_dir.mkdir(parents=True, exist_ok=True)
        dest_path = dest_dir / path.name

        if not self.dry:
            if self.mode == "copy":
                shutil.copy2(path, dest_path)
            else:
                shutil.move(path, dest_path)
                if win32com:
                    self._lnk(path.with_suffix(".lnk"), dest_path)

        revision = self._revision(path.name)
        category, tags = self._categorise(path, abbrev)

        self.rows.append(
            [
                abbrev,
                path.stem,
                path.suffix.lstrip(".").lower(),
                revision,
                category,
                ", ".join(tags),
                round(path.stat().st_size / 1048576, 2),
                datetime.fromtimestamp(path.stat().st_mtime).strftime(
                    "%Y-%m-%d %H:%M:%S"
                ),
                sha,
                f'=HYPERLINK("{dest_path.as_posix()}", "open")',
            ]
        )

    # ──────────────── helpers ──────────────────────────────────
    @staticmethod
    def _sha256(p: Path) -> str:
        h = hashlib.sha256()
        with p.open("rb") as f:
            while chunk := f.read(CHUNK):
                h.update(chunk)
        return h.hexdigest()

    @staticmethod
    def _lnk(link: Path, target: Path) -> None:
        shell = win32com.client.Dispatch("WScript.Shell")  # type: ignore
        s = shell.CreateShortCut(str(link))
        s.Targetpath = str(target)
        s.WorkingDirectory = str(target.parent)
        s.save()

    @staticmethod
    def _revision(fname: str) -> str:
        m = REV_PAT.search(fname)
        return (m["rev"].upper()) if m else ""

    # ––– categorisation –––
    def _categorise(self, p: Path, abbrev: str) -> tuple[str, list[str]]:
        name = p.stem
        for cat, val in self.rules.get("rules", {}).items():
            for trig in val.get("triggers", []):
                if trig.lower() in name.lower():
                    return cat, [trig]
        # OCR / PDF text
        text = ""
        try:
            if p.suffix.lower() == ".pdf" and pdf_text:
                text = pdf_text(str(p), maxpages=1)
            elif pytesseract and p.suffix.lower() in {".jpg", ".png", ".tif", ".tiff"}:
                text = pytesseract.image_to_string(Image.open(p))  # type: ignore
        except Exception:
            pass
        if text:
            for cat, val in self.rules.get("rules", {}).items():
                for trig in val.get("triggers", []):
                    if trig.lower() in text.lower():
                        return cat, [trig]
        if self.use_llm:
            return "LLM-TODO", []
        return self.rules.get("fallback_category", "Unknown"), []

    # ––– write excel –––
    def _write_index(self) -> None:
        out = self.tgt_root / "INDEX.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Spec Index"
        ws.append(INDEX_HEADERS)
        for r in self.rows:
            ws.append(r)
        ws.auto_filter.ref = f"A1:{get_column_letter(len(INDEX_HEADERS))}1"
        for col in range(1, len(INDEX_HEADERS) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        wb.save(out)
        LOG.info("📄 INDEX.xlsx written with %d entries", len(self.rows))


# ═════════════════════════ Qt GUI ════════════════════════════════
class SyncThread(QThread):
    status = Signal(dict)

    def __init__(self, cfg: dict):
        super().__init__()
        self.cfg = cfg
        self.worker: Optional[SpecSync] = None

    def run(self) -> None:
        self.worker = SpecSync(**self.cfg, progress_cb=self.status.emit)
        self.worker.run()

    def stop(self) -> None:
        if self.worker:
            self.worker.cancel.set()


class MainWin(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("Spec Sync v1.2")
        self.resize(880, 600)
        self.thread: Optional[SyncThread] = None

        # ––– widgets –––
        self.src_list = QListWidget()
        btn_add = QPushButton("➕ Add Source")
        btn_del = QPushButton("🗑 Remove")

        self.tgt_edit = QLineEdit()
        btn_tgt = QPushButton("Select Target")

        self.abbr_edit = QLineEdit()
        btn_abbr = QPushButton("abbreviations.txt")

        self.journal_edit = QLineEdit()
        btn_journal = QPushButton("Manual INDEX.xlsx")

        self.case_chk = QCheckBox("Case-sensitive match")
        self.mode_copy = QRadioButton("Copy")
        self.mode_move = QRadioButton("Move & Link")
        self.mode_copy.setChecked(True)
        self.dry_chk = QCheckBox("Dry-run")
        self.llm_chk = QCheckBox("Use LLM")

        self.ext_filter = QLineEdit("pdf,tif,tiff,jpg,png,gif")  # new
        self.btn_start = QPushButton("▶ Start")
        self.btn_cancel = QPushButton("⏹ Cancel")
        self.prog = QProgressBar()

        self.log_box = QTextEdit(readOnly=True)
        self.status = QStatusBar()
        self.setStatusBar(self.status)

        # ––– layout –––
        left = QVBoxLayout()
        left.addWidget(QLabel("Source Roots"))
        left.addWidget(self.src_list)
        left.addWidget(btn_add)
        left.addWidget(btn_del)

        right = QVBoxLayout()
        right.addWidget(QLabel("Target Root"))
        right.addWidget(self.tgt_edit)
        right.addWidget(btn_tgt)
        right.addWidget(QLabel("Abbreviations.txt"))
        right.addWidget(self.abbr_edit)
        right.addWidget(btn_abbr)
        right.addWidget(QLabel("Manual INDEX.xlsx (optional)"))
        right.addWidget(self.journal_edit)
        right.addWidget(btn_journal)
        right.addWidget(self.case_chk)
        right.addWidget(self.mode_copy)
        right.addWidget(self.mode_move)
        right.addWidget(self.dry_chk)
        right.addWidget(self.llm_chk)
        right.addWidget(QLabel("Extensions filter (comma)"))
        right.addWidget(self.ext_filter)
        right.addWidget(self.prog)
        btn_row = QHBoxLayout()
        btn_row.addWidget(self.btn_start)
        btn_row.addWidget(self.btn_cancel)
        right.addLayout(btn_row)

        main = QHBoxLayout()
        main.addLayout(left, 2)
        main.addLayout(right, 3)

        main_tab = QWidget()         # ← оборачиваем layout в отдельный QWidget
        main_tab.setLayout(main)

        log_tab = QTabWidget()
        log_tab.addTab(main_tab, "Sync")   # ← нужен заголовок!
        log_tab.addTab(self.log_box, "Log")

        central = QVBoxLayout()
        central.addWidget(log_tab)

        w = QWidget()
        w.setLayout(central)
        self.setCentralWidget(w)

        # ––– menu –––
        m_tools = self.menuBar().addMenu("&Tools")
        act_dump = QAction("Create Dump (zip)", self)
        act_restore = QAction("Restore Dump", self)
        act_open_idx = QAction("Open INDEX.xlsx", self)
        m_tools.addActions([act_dump, act_restore, act_open_idx])

        # ––– connections –––
        btn_add.clicked.connect(self._src_add)
        btn_del.clicked.connect(lambda: self._rm_selected(self.src_list))
        btn_tgt.clicked.connect(lambda: self._pick_folder(self.tgt_edit))
        btn_abbr.clicked.connect(lambda: self._pick_file(self.abbr_edit, "*.txt"))
        btn_journal.clicked.connect(lambda: self._pick_file(self.journal_edit, "*.xlsx"))
        self.btn_start.clicked.connect(self._start)
        self.btn_cancel.clicked.connect(self._cancel)
        act_dump.triggered.connect(self._dump)
        act_restore.triggered.connect(self._restore)
        act_open_idx.triggered.connect(self._open_index)

    # ––– ui helpers –––
    def _src_add(self) -> None:
        d = QFileDialog.getExistingDirectory(self, "Choose Source folder")
        if d:
            self.src_list.addItem(QListWidgetItem(d))

    def _rm_selected(self, lw: QListWidget) -> None:
        for idx in lw.selectedIndexes():
            lw.takeItem(idx.row())

    def _pick_folder(self, edit: QLineEdit) -> None:
        d = QFileDialog.getExistingDirectory(self, "Pick folder")
        if d:
            edit.setText(d)

    def _pick_file(self, edit: QLineEdit, flt: str) -> None:
        f, _ = QFileDialog.getOpenFileName(self, "Pick file", filter=flt)
        if f:
            edit.setText(f)

    # ––– main workflow –––
    def _cfg(self) -> dict | None:
        srcs = [Path(self.src_list.item(i).text()) for i in range(self.src_list.count())]
        if not srcs:
            QMessageBox.warning(self, "Need sources", "Add at least one source!")
            return None
        tgt = Path(self.tgt_edit.text())
        if not tgt.exists():
            QMessageBox.warning(self, "Target", "Valid target folder required")
            return None
        abbr = Path(self.abbr_edit.text())
        if not abbr.exists():
            QMessageBox.warning(self, "Abbrev", "Select abbreviations.txt")
            return None

        exts = {f".{e.strip().lower()}" for e in self.ext_filter.text().split(",") if e.strip()}
        if exts:
            global ALLOWED_EXTS  # override runtime
            ALLOWED_EXTS = exts

        cfg = dict(
            source_roots=srcs,
            target_root=tgt,
            abbreviations_file=abbr,
            rules_file=Path("rules.yaml") if Path("rules.yaml").exists() else None,
            journal_file=Path(self.journal_edit.text()) if self.journal_edit.text() else None,
            mode="move" if self.mode_move.isChecked() else "copy",
            case_sensitive=self.case_chk.isChecked(),
            use_llm=self.llm_chk.isChecked(),
            dry_run=self.dry_chk.isChecked(),
        )
        return cfg

    def _start(self) -> None:
        if self.thread and self.thread.isRunning():
            QMessageBox.warning(self, "Busy", "Sync already running.")
            return
        cfg = self._cfg()
        if not cfg:
            return
        self.prog.setRange(0, 0)
        self.log_box.clear()
        self.thread = SyncThread(cfg)
        self.thread.status.connect(self._status)
        self.thread.finished.connect(self._done)
        self.thread.start()

    def _cancel(self) -> None:
        if self.thread:
            self.thread.stop()

    # ––– signals –––
    @Slot(dict)
    def _status(self, st: dict) -> None:
        msg = (
            f'Scanned {st["scanned"]}  '
            f'Unique {st["unique"]}  Dup {st["dup"]}  '
            f'Skipped {st["skipped"]}  Err {st["errors"]}'
        )
        self.status.showMessage(msg)
        self.log_box.append(msg)

    def _done(self) -> None:
        self.prog.setRange(0, 1)
        QMessageBox.information(self, "Done", "Sync finished")

    # ––– tools –––
    def _dump(self) -> None:
        tgt = Path(self.tgt_edit.text())
        if not tgt.exists():
            return
        dump, _ = QFileDialog.getSaveFileName(self, "Dump path", "specs_dump.zip")
        if not dump:
            return
        with zipfile.ZipFile(dump, "w", zipfile.ZIP_DEFLATED) as z:
            for p in tgt.rglob("*"):
                z.write(p, p.relative_to(tgt))
        QMessageBox.information(self, "Dump", "ZIP created")

    def _restore(self) -> None:
        tgt = Path(self.tgt_edit.text())
        if not tgt.exists():
            return
        dump, _ = QFileDialog.getOpenFileName(self, "Select dump", filter="*.zip")
        if not dump:
            return
        with zipfile.ZipFile(dump) as z:
            z.extractall(tgt)
        QMessageBox.information(self, "Restore", "Dump restored")

    def _open_index(self) -> None:
        idx = Path(self.tgt_edit.text()) / "INDEX.xlsx"
        if idx.exists():
            os.startfile(idx)  # Windows


# ═════════════════════════ CLI entry ═════════════════════════════
def cli() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("source", nargs="+", type=Path)
    ap.add_argument("-t", "--target", required=True, type=Path)
    ap.add_argument("-a", "--abbrev", required=True, type=Path)
    ap.add_argument("-j", "--journal", type=Path)
    ap.add_argument("-m", "--mode", choices=["copy", "move"], default="copy")
    ap.add_argument("--ci", help="case-insensitive", action="store_true")
    ap.add_argument("--dry", action="store_true")
    args = ap.parse_args()

    ss = SpecSync(
        source_roots=args.source,
        target_root=args.target,
        abbreviations_file=args.abbrev,
        journal_file=args.journal,
        mode=args.mode,
        case_sensitive=not args.ci,
        dry_run=args.dry,
    )
    ss.run()


if __name__ == "__main__":
    if sys.stdin.isatty():  # launched by double-click
        app = QApplication(sys.argv)
        MainWin().show()
        sys.exit(app.exec())
    cli()

# ────────────────────────── CHANGELOG ────────────────────────────
# v1.2 (2025-05-26)
#   • Аббревиатуры длиной 1-2 символа + знаки +/-
#   • Сверка с Target-папкой и ручным INDEX.xlsx
#   • Фильтр расширений в GUI
#   • Открыть INDEX.xlsx из меню
#   • Улучшенные логи, stats, colours
