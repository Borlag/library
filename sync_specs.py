# =============================  sync_specs.py  =============================
"""
sync_specs.py – Aircraft-spec repository synchroniser  (GUI = PySide 6)
------------------------------------------------------------------------
Back-end logic identical to the earlier version (see doc-string there);
only the GUI layer was re-implemented with PySide 6 at the user’s
request.  Run `python sync_specs.py` to launch the GUI, or pass CLI args
to work headless.
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import logging
import os
import re
import shutil
import sqlite3
import sys
import threading
import time
from contextlib import contextmanager
from dataclasses import dataclass
from email.message import EmailMessage
from pathlib import Path
from typing import Callable, Iterator, Optional

# ----------  third-party  ----------
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import yaml

# --- PySide 6 -------------------------------------------------------------
from PySide6.QtCore import (Qt, QThread, Signal, QObject)
from PySide6.QtGui  import QAction
from PySide6.QtWidgets import (
    QApplication, QFileDialog, QGridLayout, QHBoxLayout, QLabel,
    QLineEdit, QMainWindow, QMenu, QMenuBar, QMessageBox,  QPushButton,
    QProgressBar, QRadioButton, QStatusBar, QTextEdit, QVBoxLayout, QWidget,
    QCheckBox
)

# pywin32 only on Windows for .lnk shortcuts -- import guarded
if os.name == "nt":
    import pythoncom  # noqa: F401
    import win32com.client  # type: ignore


# ---------------------------  GLOBAL CONSTANTS  --------------------------- #
SCAN_EXTS = {".pdf"}
ABBREV_RE = re.compile(r"[A-Z0-9\-]{3,15}")
REV_RE = re.compile(r"(?:[_\- ]REV[_\-\.\s]?|[_\- ]R)([A-Z])", re.IGNORECASE)
INDEX_COLUMNS = [
    "Abbrev",
    "File_Name",
    "Revision",
    "Category",
    "Tags",
    "Size_MB",
    "Modified",
    "SHA256",
    "Link",
]
DB_SCHEMA = """
CREATE TABLE IF NOT EXISTS files (
    sha256 TEXT PRIMARY KEY,
    abbrev TEXT,
    path TEXT,
    revision TEXT,
    category TEXT,
    tags TEXT,
    size_mb REAL,
    modified TEXT
);
"""


# -----------------------------  DATA CLASSES  ----------------------------- #
@dataclass
class FileMeta:
    abbrev: str
    file_name: str
    revision: str
    category: str
    tags: str
    size_mb: float
    modified: str
    sha256: str
    link: str

    def as_row(self) -> list:
        return [
            self.abbrev,
            self.file_name,
            self.revision,
            self.category,
            self.tags,
            f"{self.size_mb:.2f}",
            self.modified,
            self.sha256,
            self.link,
        ]


# ---------------------------  UTILITY FUNCTIONS  -------------------------- #
def sha256sum(path: Path, buf_size: int = 1 << 18) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        while chunk := f.read(buf_size):
            h.update(chunk)
    return h.hexdigest()

def build_abbrev_regex(abbrevs: set[str]) -> re.Pattern:
    """
    Вернуть компилированный regex, который ищет *любое* из переданных
    аббревиатур (без игнорирования регистра).  Сортировка по длине ↓
    гарантирует, что найдётся самое длинное совпадение.
    """
    escaped = sorted((re.escape(a) for a in abbrevs), key=len, reverse=True)
    pattern = "|".join(escaped)
    return re.compile(pattern)


def extract_abbrev(name: str, pattern: re.Pattern) -> Optional[str]:
    """
    Найти первое точное совпадение whitelist-аббревиатуры *в оригинальном
    регистре* имени файла.  Возвращает аббревиатуру или None.
    """
    m = pattern.search(name)
    return m.group(0) if m else None


def extract_revision(name: str) -> str:
    m = REV_RE.search(name)
    return m.group(1).upper() if m else ""


def create_shortcut(src: Path, dst: Path) -> None:
    shell = win32com.client.Dispatch("WScript.Shell")
    shortcut = shell.CreateShortcut(str(dst))
    shortcut.TargetPath = str(src)
    shortcut.WorkingDirectory = str(src.parent)
    shortcut.IconLocation = str(src)
    shortcut.save()


def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


@contextmanager
def sqlite_conn(db_path: Path) -> Iterator[sqlite3.Connection]:
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA journal_mode=WAL;")
    try:
        yield conn
    finally:
        conn.commit()
        conn.close()


def send_error_email(subject: str, body: str) -> None:
    load_dotenv()
    smtp_host = os.getenv("SMTP_HOST")
    smtp_user = os.getenv("SMTP_USER")
    smtp_pass = os.getenv("SMTP_PASS")
    to_addr = os.getenv("SMTP_TO")
    if not all((smtp_host, smtp_user, smtp_pass, to_addr)):
        logging.warning("SMTP not configured – e-mail skipped.")
        return
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = smtp_user
    msg["To"] = to_addr
    msg.set_content(body)
    import smtplib

    with smtplib.SMTP_SSL(smtp_host) as s:
        s.login(smtp_user, smtp_pass)
        s.send_message(msg)


# ---------------------------  CORE SYNC ENGINE  --------------------------- #
class SpecSync:
    """Synchronise unstructured spec tree → clean repository."""

    def __init__(  # часть аргументов опущена для краткости
        self,
        source_root: Path,
        target_root: Path,
        abbreviations: Path,
        rules_yaml: Path,
        mode: str = "copy",
        dry_run: bool = False,
        use_llm: bool = False,
        progress_cb: Optional[Callable[[str], None]] = None,
    ):
        self.src_root = source_root
        self.tgt_root = target_root
        self.mode = mode
        self.dry_run = dry_run
        self.use_llm = use_llm
        self.progress = progress_cb or (lambda m: None)

        with abbreviations.open(encoding="utf-8") as f:
            self.whitelist: set[str] = {ln.strip() for ln in f if ln.strip()}
        self.abbrev_regex = build_abbrev_regex(self.whitelist)

        with rules_yaml.open(encoding="utf-8") as f:
            self.rules = yaml.safe_load(f) or {}

        self.db_path = target_root / "inventory.db"
        ensure_dir(target_root)
        with sqlite_conn(self.db_path) as conn:
            conn.executescript(DB_SCHEMA)

        self.scanned = self.unique = self.duplicates = self.errors = 0

    # ----------  PUBLIC API  ---------- #
    def run(self) -> None:
        try:
            self._sync()
        except Exception as exc:  # noqa: BLE001
            logging.exception("Fatal error")
            send_error_email("Spec Sync FAILED", str(exc))
            raise

    # ----------  INTERNAL  ---------- #
    def _sync(self) -> None:
        pdf_files = list(self._iter_files())
        total = len(pdf_files)
        logging.info("Found %d candidate files", total)

        index_path = self.tgt_root / "INDEX.xlsx"
        wb, ws = self._prepare_index(index_path)

        for n, path in enumerate(pdf_files, 1):
            self.progress(f"Scanning ({n}/{total}) – {path.name}")
            try:
                self._process_file(path, ws)
            except Exception:
                self.errors += 1
                logging.exception("Error processing %s", path)
            time.sleep(0)  # yield thread

        wb.save(index_path)
        self._autosize_columns(index_path)
        self.progress(
            f"Done. Scanned:{self.scanned}  Unique:{self.unique}  "
            f"Duplicates:{self.duplicates}  Errors:{self.errors}"
        )

    def _iter_files(self) -> Iterator[Path]:
        for p in self.src_root.rglob("*"):
            if p.is_file() and p.suffix.lower() in SCAN_EXTS:
                self.scanned += 1
                if extract_abbrev(p.name, self.abbrev_regex):    # ### FIX
                    yield p

    # -- INDEX helpers -----------------------------------------------------
    def _prepare_index(self, path: Path):
        if path.exists():
            wb = load_workbook(path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(INDEX_COLUMNS)
            ws.auto_filter.ref = f"A1:{get_column_letter(len(INDEX_COLUMNS))}1"
        return wb, ws

    def _autosize_columns(self, path: Path) -> None:
        wb = load_workbook(path)
        ws = wb.active
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = max(len(str(cell.value or "")) for cell in col) + 2
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 80)
        wb.save(path)

    # -- per-file processing ----------------------------------------------
    def _process_file(self, src: Path, ws) -> None:
        abbrev = extract_abbrev(src.name, self.abbrev_regex)     # ### FIX
        if not abbrev:
            return

        sha = sha256sum(src)
        with sqlite_conn(self.db_path) as conn:
            cur = conn.cursor()
            cur.execute("SELECT path FROM files WHERE sha256=?", (sha,))
            row = cur.fetchone()

        if row:
            self.duplicates += 1
            if self.mode == "move_link" and not self.dry_run:
                self._make_shortcut(src, Path(row[0]))
            return

        self.unique += 1
        revision = extract_revision(src.stem)
        category, tags = self._categorise(src)

        dest_dir = self.tgt_root / abbrev
        ensure_dir(dest_dir)
        dest = self._unique_dest(dest_dir, src.name)

        link_formula = f'=HYPERLINK("{dest.as_posix()}", "{dest.name}")'

        meta = FileMeta(
            abbrev=abbrev,
            file_name=dest.name,
            revision=revision,
            category=category,
            tags=tags,
            size_mb=src.stat().st_size / (1024 ** 2),
            modified = dt.datetime.fromtimestamp(src.stat().st_mtime,dt.timezone.utc).isoformat(),
            sha256=sha,
            link=link_formula,
        )

        if not self.dry_run:
            if self.mode == "copy":
                shutil.copy2(src, dest)
            else:
                shutil.move(src, dest)
                self._make_shortcut(src, dest)

        with sqlite_conn(self.db_path) as conn:
            conn.execute(
                "INSERT OR REPLACE INTO files VALUES (?,?,?,?,?,?,?,?)",
                (
                    sha,
                    abbrev,
                    str(dest),
                    revision,
                    category,
                    tags,
                    meta.size_mb,
                    meta.modified,
                ),
            )
        ws.append(meta.as_row())

    def _unique_dest(self, dest_dir: Path, filename: str) -> Path:
        dest = dest_dir / filename
        if not dest.exists():
            return dest
        base, ext = os.path.splitext(filename)
        i = 1
        while (dest_dir / f"{base}__{i}{ext}").exists():
            i += 1
        return dest_dir / f"{base}__{i}{ext}"

    def _make_shortcut(self, original: Path, target: Path) -> None:
        lnk = original.with_suffix(".lnk")
        if not lnk.exists() and os.name == "nt":
            create_shortcut(target, lnk)

    # -- categorisation ----------------------------------------------------
    def _categorise(self, path: Path) -> tuple[str, str]:
        text = path.name.lower()
        for cat, kw_list in self.rules.items():
            if any(kw.lower() in text for kw in kw_list):
                return cat, ",".join(kw_list)
        if self.use_llm:
            return self._llm_categorise(text)
        return "Uncategorised", ""

    def _llm_categorise(self, text: str) -> tuple[str, str]:
        return "LLM-Unk", ""  # TODO: plug real call


# -----------------------  GUI (PySide 6)  ---------------------------------- #
class Worker(QThread):
    progress = Signal(str)
    finished = Signal()

    def __init__(self, engine: SpecSync):
        super().__init__()
        self.engine = engine

    def run(self) -> None:
        self.engine.progress = self.progress.emit
        try:
            self.engine.run()
        finally:
            self.finished.emit()


class MainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("Spec Repository Sync")
        self.setMinimumSize(900, 600)
        self._build_ui()
        self.worker: Optional[Worker] = None
        logging.basicConfig(filename="sync.log",
                            level=logging.INFO,
                            format="%(asctime)s %(levelname)s %(message)s")

    # ------------------  UI LAYOUT  ------------------ #
    def _build_ui(self) -> None:
        central = QWidget(self)
        self.setCentralWidget(central)
        grid = QGridLayout()
        central.setLayout(grid)

        # Inputs
        self.src_edit = QLineEdit()
        self.dst_edit = QLineEdit()
        self.abbr_edit = QLineEdit()
        self.rules_edit = QLineEdit("rules.yaml")
        src_btn = QPushButton("Browse…")
        dst_btn = QPushButton("Browse…")
        abbr_btn = QPushButton("Browse…")
        rules_btn = QPushButton("Browse…")

        src_btn.clicked.connect(lambda: self._choose_dir(self.src_edit))
        dst_btn.clicked.connect(lambda: self._choose_dir(self.dst_edit))
        abbr_btn.clicked.connect(lambda: self._choose_file(self.abbr_edit))
        rules_btn.clicked.connect(lambda: self._choose_file(self.rules_edit))

        grid.addWidget(QLabel("Source Root"), 0, 0)
        grid.addWidget(self.src_edit, 0, 1)
        grid.addWidget(src_btn, 0, 2)

        grid.addWidget(QLabel("Target Root"), 1, 0)
        grid.addWidget(self.dst_edit, 1, 1)
        grid.addWidget(dst_btn, 1, 2)

        grid.addWidget(QLabel("abbreviations.txt"), 2, 0)
        grid.addWidget(self.abbr_edit, 2, 1)
        grid.addWidget(abbr_btn, 2, 2)

        grid.addWidget(QLabel("rules.yaml"), 3, 0)
        grid.addWidget(self.rules_edit, 3, 1)
        grid.addWidget(rules_btn, 3, 2)

        # Mode radio buttons
        self.copy_radio = QRadioButton("Copy Mode")
        self.move_radio = QRadioButton("Move & Link Mode")
        self.copy_radio.setChecked(True)

        mode_box = QHBoxLayout()
        mode_box.addWidget(self.copy_radio)
        mode_box.addWidget(self.move_radio)
        grid.addLayout(mode_box, 4, 0, 1, 3)

        # Checkboxes
        self.llm_chk = QCheckBox("Use LLM fallback")
        self.dry_chk = QCheckBox("Dry-Run")
        chk_box = QHBoxLayout()
        chk_box.addWidget(self.llm_chk)
        chk_box.addWidget(self.dry_chk)
        grid.addLayout(chk_box, 5, 0, 1, 3)

        # Progress + log
        self.progress = QProgressBar()
        self.progress.setRange(0, 0)        # marquee style
        self.progress.hide()

        self.log_edit = QTextEdit()
        self.log_edit.setReadOnly(True)

        grid.addWidget(self.progress, 6, 0, 1, 3)
        grid.addWidget(self.log_edit, 7, 0, 1, 3)

        # Buttons
        self.start_btn = QPushButton("Start")
        self.cancel_btn = QPushButton("Cancel")
        self.exit_btn = QPushButton("Exit")
        self.cancel_btn.setEnabled(False)

        btn_box = QHBoxLayout()
        btn_box.addWidget(self.start_btn)
        btn_box.addWidget(self.cancel_btn)
        btn_box.addWidget(self.exit_btn)
        grid.addLayout(btn_box, 8, 0, 1, 3)

        self.start_btn.clicked.connect(self._start)
        self.cancel_btn.clicked.connect(self._cancel)
        self.exit_btn.clicked.connect(self.close)

        # Status bar
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)

        # Menu
        menu = QMenuBar()
        tools_menu = QMenu("Tools", self)
        safety_menu = tools_menu.addMenu("Safety")
        act_dump = QAction("Create Dump", self)
        act_restore = QAction("Restore from Dump", self)
        safety_menu.addAction(act_dump)
        safety_menu.addAction(act_restore)
        menu.addMenu(tools_menu)
        self.setMenuBar(menu)

        act_dump.triggered.connect(self._create_dump)
        act_restore.triggered.connect(self._restore_dump)

    # ------------------  EVENTS  ------------------ #
    def _choose_dir(self, target_edit: QLineEdit) -> None:
        directory = QFileDialog.getExistingDirectory(self, "Choose folder")
        if directory:
            target_edit.setText(directory)

    def _choose_file(self, target_edit: QLineEdit) -> None:
        file_, _ = QFileDialog.getOpenFileName(self, "Choose file", filter="All (*.*)")
        if file_:
            target_edit.setText(file_)

    def _start(self) -> None:
        if self.worker and self.worker.isRunning():
            QMessageBox.warning(self, "Sync", "Sync already running.")
            return

        src = Path(self.src_edit.text())
        dst = Path(self.dst_edit.text())
        abbr = Path(self.abbr_edit.text())
        rules = Path(self.rules_edit.text())

        if not (src.is_dir() and dst and abbr.exists() and rules.exists()):
            QMessageBox.critical(self, "Sync", "Please select valid paths.")
            return

        mode = "copy" if self.copy_radio.isChecked() else "move_link"

        engine = SpecSync(
            source_root=src,
            target_root=dst,
            abbreviations=abbr,
            rules_yaml=rules,
            mode=mode,
            dry_run=self.dry_chk.isChecked(),
            use_llm=self.llm_chk.isChecked(),
        )

        self.worker = Worker(engine)
        self.worker.progress.connect(self._log)
        self.worker.finished.connect(self._finished)
        self.worker.start()

        self.start_btn.setEnabled(False)
        self.cancel_btn.setEnabled(True)
        self.progress.show()
        self.log_edit.clear()
        self._log("== SYNC STARTED ==")

    def _cancel(self) -> None:
        if self.worker and self.worker.isRunning():
            self.worker.terminate()
            self.worker.wait()
        self._finished()

    def _finished(self) -> None:
        self.progress.hide()
        self.start_btn.setEnabled(True)
        self.cancel_btn.setEnabled(False)
        self._log("== SYNC FINISHED ==")

    def _log(self, message: str) -> None:
        ts = time.strftime("%H:%M:%S")
        self.log_edit.append(f"[{ts}] {message}")
        self.status_bar.showMessage(message)
        logging.info(message)

    # ------------------  SAFETY  ------------------ #
    def _create_dump(self) -> None:
        dst_root = Path(self.dst_edit.text())
        db = dst_root / "inventory.db"
        if not db.exists():
            QMessageBox.warning(self, "Dump", "No inventory.db found.")
            return
        ts = dt.datetime.now().strftime("%Y%m%d")
        dump_file = dst_root / f"dump_{ts}.json"
        try:
            with sqlite_conn(db) as conn, dump_file.open("w", encoding="utf-8") as f:
                rows = conn.execute("SELECT * FROM files").fetchall()
                cols = [d[0] for d in conn.execute("PRAGMA table_info(files)")]
                json.dump([dict(zip(cols, r)) for r in rows], f, indent=2)
            QMessageBox.information(self, "Dump", f"Dump created:\n{dump_file}")
        except Exception as e:  # noqa: BLE001
            QMessageBox.critical(self, "Dump", f"Dump failed:\n{e}")

    def _restore_dump(self) -> None:
        dump_path, _ = QFileDialog.getOpenFileName(self, "Select dump_*.json", filter="JSON (*.json)")
        if not dump_path:
            return
        dst_root = Path(self.dst_edit.text())
        db = dst_root / "inventory.db"
        try:
            with sqlite_conn(db) as conn, open(dump_path, encoding="utf-8") as f:
                data = json.load(f)
                conn.execute("DELETE FROM files")
                for row in data:
                    conn.execute(
                        "INSERT INTO files VALUES (?,?,?,?,?,?,?,?)",
                        (
                            row["sha256"],
                            row["abbrev"],
                            row["path"],
                            row["revision"],
                            row["category"],
                            row["tags"],
                            row["size_mb"],
                            row["modified"],
                        ),
                    )
            QMessageBox.information(self, "Restore", "Inventory restored from dump.")
        except Exception as e:  # noqa: BLE001
            QMessageBox.critical(self, "Restore", f"Restore failed:\n{e}")


# ---------------------------  CLI ENTRY POINT  ---------------------------- #
def cli() -> None:
    parser = argparse.ArgumentParser(description="Synchronise spec repository (headless)")
    parser.add_argument("source", type=Path)
    parser.add_argument("target", type=Path)
    parser.add_argument("--abbr", type=Path, default="abbreviations.txt")
    parser.add_argument("--rules", type=Path, default="rules.yaml")
    parser.add_argument("--mode", choices=("copy", "move_link"), default="copy")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--use-llm", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(filename="sync.log",
                        level=logging.INFO,
                        format="%(asctime)s %(levelname)s %(message)s")
    SpecSync(
        source_root=args.source,
        target_root=args.target,
        abbreviations=args.abbr,
        rules_yaml=args.rules,
        mode=args.mode,
        dry_run=args.dry_run,
        use_llm=args.use_llm,
    ).run()


if __name__ == "__main__":
    if len(sys.argv) > 1 and Path(sys.argv[1]).exists():
        cli()
    else:
        app = QApplication(sys.argv)
        w = MainWindow()
        w.show()
        sys.exit(app.exec())

# --------------------------------------------------------------------------
