import os
import csv
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional

import openpyxl
from openpyxl.styles import Font

from models import ProcessorConfig
from renaming_rules import FileNameFormatter
from utils import scan_all_files_recursive, looks_like_office_temp

class ExcelCatalog:
    """
    Генерация/обновление Excel-каталога.
    - "Abbrev"     — распознанная аббревиатура (папка).
    - "File_Name"  — display name (нормализованное название стандарта без расширения, UPPERCASE).
    - "Format"     — фактическое расширение OS-файла (pdf/doc/...).
    - "Revision"   — попытка извлечь ревизию (REV/буква/год/выпуск).
    - "Category Tags" — импортируемые теги по нормализованному ключу.
    - "Size_MB" / "Modified" / "SHA256" (пока не считаем — оставляем пусто).
    - "File_Link"  — гиперссылка на реальный путь OS-файла.

    Режим "append" не удаляет пользовательские изменения в существующей книге:
    сравнение по (link_target, modified_text).
    """

    HEADERS = ["Abbrev", "File_Name", "Format", "Revision", "Category Tags",
               "Size_MB", "Modified", "SHA256", "File_Link"]

    def __init__(self, config: ProcessorConfig, log_fn, parent_logger, abbrev_fn):
        self.config = config
        self.log = log_fn
        self.get_abbrev_fn = abbrev_fn
        self.logger = parent_logger.getChild(self.__class__.__name__)
        self.xlsx_path = os.path.join(self.config.destination_folder, self.config.catalog_filename)
        self.formatter = FileNameFormatter(force_uppercase=self.config.force_uppercase_names)

    def _normalize_key(self, s: str) -> str:
        if not s:
            return ""
        base = os.path.splitext(os.path.basename(s))[0]
        base = base.replace('/', '_')
        base = base.replace('-', '').replace('_', '').replace(' ', '')
        return base.lower()

    def _load_tags_map(self) -> Dict[str, str]:
        tags: Dict[str, str] = {}
        if not self.config.tag_file:
            return tags
        try:
            if self.config.tag_file.lower().endswith(".xlsx"):
                wb = openpyxl.load_workbook(self.config.tag_file, data_only=True)
                ws = wb.active
                for row in ws.iter_rows(min_row=2):
                    name_val = (row[0].value or "").strip() if len(row) >= 1 and row[0].value else ""
                    tags_val = (row[6].value or "").strip() if len(row) >= 7 and row[6].value else ""
                    if name_val:
                        key = self._normalize_key(name_val)
                        if key:
                            tags[key] = tags_val
            elif self.config.tag_file.lower().endswith(".csv"):
                with open(self.config.tag_file, "r", encoding="utf-8", newline="") as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                    for r in rows[1:]:
                        if not r:
                            continue
                        name_val = (r[0] or "").strip() if len(r) >= 1 else ""
                        tags_val = (r[6] or "").strip() if len(r) >= 7 else ""
                        if name_val:
                            key = self._normalize_key(name_val)
                            if key:
                                tags[key] = tags_val
        except Exception as e:
            self.log(f"Не удалось импортировать теги: {e}", "WARNING")
        return tags

    def _open_or_create_wb(self):
        wb = None
        ws = None
        loaded_existing = False
        if self.config.append_mode and os.path.exists(self.xlsx_path):
            try:
                wb = openpyxl.load_workbook(self.xlsx_path)
                ws = wb.active
                loaded_existing = True
            except Exception as e:
                self.log(f"Не удалось открыть существующий каталог: {e}", "WARNING")
                wb = None
                ws = None

        if wb is None:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Каталог Спецификаций"
            ws.append(self.HEADERS)
            for cell in ws[1]:
                cell.font = Font(bold=True)

        return wb, ws, loaded_existing

    def _gather_existing_rows(self, ws):
        rows = {}
        by_path = {}
        by_norm_key = {}
        by_size_mod = {}
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            values = []
            for idx in range(len(self.HEADERS)):
                if idx < len(row):
                    values.append(row[idx].value)
                else:
                    values.append("")
            link_cell = row[8] if len(row) >= 9 else None
            link_target = link_cell.hyperlink.target if (link_cell and link_cell.hyperlink) else ""
            entry = {
                "values": values,
                "link": link_target or "",
            }

            norm_path = os.path.normcase(link_target) if link_target else ""
            if norm_path:
                by_path[norm_path] = i

            norm_display = self._normalize_key(values[1] or "")
            norm_link_name = self._normalize_key(os.path.basename(link_target)) if link_target else ""
            keys = set(filter(None, [norm_display, norm_link_name]))
            for key in keys:
                by_norm_key.setdefault(key, []).append(i)

            size_val = values[5]
            try:
                size_float = float(size_val) if size_val is not None and size_val != "" else None
            except (TypeError, ValueError):
                size_float = None
            modified_val = values[6] if len(values) > 6 else None
            modified_str = str(modified_val) if modified_val else ""
            entry["size_mb"] = size_float
            entry["modified"] = modified_str
            if size_float is not None and modified_str:
                key = (round(size_float, 3), modified_str)
                by_size_mod.setdefault(key, []).append(i)

            rows[i] = entry

        return rows, by_path, by_norm_key, by_size_mod

    def _find_existing_match(self, filepath, norm_disp, norm_raw, size_mb, modified,
                              by_path, by_norm_key, by_size_mod, existing_rows, matched_rows):
        norm_path = os.path.normcase(filepath)
        idx = by_path.get(norm_path)
        if idx and idx not in matched_rows:
            return idx

        for key in filter(None, [norm_disp, norm_raw]):
            candidates = by_norm_key.get(key)
            if not candidates:
                continue
            for cand in candidates:
                if cand not in matched_rows:
                    return cand

        if size_mb is not None and modified:
            key = (round(size_mb, 3), modified)
            candidates = by_size_mod.get(key)
            if candidates:
                for cand in candidates:
                    if cand not in matched_rows:
                        return cand

        return None

    # --- Извлечение ревизии ---
    def _extract_revision(self, display_name: str, filename_no_ext: str) -> str:
        """
        Расширенное извлечение ревизии.
        Поддерживаем распространённые варианты:
        - REV / REVISION / REV. / REV: / REVISION: + суффикс
        - ISSUE / ED(ITION) / CHG / CHANGE / AMDT / AMENDMENT / MOD + суффикс
        - BACR15GA_REV.K -> "K" (учёт точек/подчёркиваний)
        - Окончание '_X' или '-X' для QPL/REV имен (например, "BMS10-103 QPL_D")
        - AMSXXXX/XX -> "XX"
        - Одиночная буква на конце обозначения ("...-416F") как запасной вариант
        - Год (19xx/20xx) как последний fallback
        """

        contexts = [filename_no_ext.upper(), display_name.upper()]
        keyword_patterns = [
            r'\bREV(?:ISION)?[\s._-]*([A-Z0-9]+)\b',
            r'\bREV\.[\s._-]*([A-Z0-9]+)\b',
            r'\bREV:[\s._-]*([A-Z0-9]+)\b',
            r'\bISSUE[\s._-]*([A-Z0-9]+)\b',
            r'\bED(?:ITION)?[\s._-]*([A-Z0-9]+)\b',
            r'\bCHG[\s._-]*([A-Z0-9]+)\b',
            r'\bCHANGE[\s._-]*([A-Z0-9]+)\b',
            r'\bAMDT[\s._-]*([A-Z0-9]+)\b',
            r'\bAMEND(?:MENT)?[\s._-]*([A-Z0-9]+)\b',
            r'\bMOD(?:IFICATION)?[\s._-]*([A-Z0-9]+)\b',
        ]

        for ctx in contexts:
            for pat in keyword_patterns:
                m = re.search(pat, ctx)
                if m:
                    return m.group(1).strip('-_.')

        s = display_name.upper()
        fn = filename_no_ext.upper()

        # AMS: '/N' или '_N' в конце display
        m = re.search(r'/([A-Z0-9]{1,2})$', s)
        if m:
            return m.group(1)

        # QPL и подобные: окончание '_X' или '-X' после упоминания REV/QPL
        tail_candidates = [fn, s]
        for ctx in tail_candidates:
            tail = re.search(r'[_-]([A-Z]{1,3})$', ctx)
            if tail and any(token in ctx for token in ('QPL', 'REV', 'ISSUE', 'ED', 'EDITION', 'CHG', 'CHANGE', 'AMDT', 'AMEND', 'MOD')):
                return tail.group(1)

        # AMS/SAE буква на конце кода "...-416F"
        m = re.search(r'-[0-9A-Z]+([A-Z])$', s)
        if m:
            return m.group(1)

        # Год (как fallback)
        m = re.search(r'\b(19\d{2}|20\d{2})\b', fn)
        if m:
            return m.group(1)

        return ""

    def generate(self):
        self.log("Начало генерации каталога.", "INFO")
        os.makedirs(self.config.destination_folder, exist_ok=True)

        wb, ws, loaded_existing = self._open_or_create_wb()
        existing_rows = {}
        by_path = {}
        by_norm_key = {}
        by_size_mod = {}
        if loaded_existing:
            existing_rows, by_path, by_norm_key, by_size_mod = self._gather_existing_rows(ws)
            self.log(f"Режим 'Дополнять': найдено {len(existing_rows)} записей.", "INFO")

        tags_map = self._load_tags_map()
        all_files = scan_all_files_recursive(self.config.destination_folder)
        link_font = Font(color="0000FF", underline="single")
        row_count_added = 0
        row_count_updated = 0
        matched_rows = set()
        updated_entries = {}
        new_entries = []

        corrupt_norm = os.path.normcase(self.config.corrupt_folder)

        for filepath, folder_name, filename in all_files:
            if looks_like_office_temp(filename, self.config.office_temp_prefix):
                continue
            if filename == self.config.catalog_filename:
                continue
            if filename.upper().startswith("AUDIT_REPORT"):
                continue
            parts_norm = [os.path.normcase(part) for part in Path(filepath).parts]
            if corrupt_norm in parts_norm:
                continue

            base, ext = os.path.splitext(filename)
            abbrev_tuple = self.get_abbrev_fn(filename)
            if abbrev_tuple and len(abbrev_tuple) >= 2:
                abbrev, orig_abbrev = abbrev_tuple[0], abbrev_tuple[1]
            else:
                abbrev, orig_abbrev = folder_name, folder_name

            display = self.formatter.format_display_name(filename, abbrev, orig_abbrev)

            try:
                stats = os.stat(filepath)
                size_mb = round(stats.st_size / (1024 * 1024), 3)
                modified = datetime.fromtimestamp(stats.st_mtime).strftime('%Y-%m-%d %H:%M')
            except FileNotFoundError:
                size_mb, modified = 0, ""

            norm_disp = self._normalize_key(display)
            norm_raw = self._normalize_key(filename)
            tags_value = tags_map.get(norm_disp) or tags_map.get(norm_raw) or ""

            revision = self._extract_revision(display, base)

            match_idx = self._find_existing_match(
                filepath,
                norm_disp,
                norm_raw,
                size_mb,
                modified,
                by_path,
                by_norm_key,
                by_size_mod,
                existing_rows,
                matched_rows,
            ) if existing_rows else None

            row_values = [
                abbrev,
                display,
                ext.lstrip('.').lower(),
                revision,
                tags_value,
                size_mb,
                modified,
                "",
                "Открыть файл"
            ]
            if match_idx:
                existing_entry = existing_rows.get(match_idx, {})
                existing_values = existing_entry.get("values", [])
                if existing_values:
                    if not row_values[4] and len(existing_values) > 4 and existing_values[4]:
                        row_values[4] = existing_values[4]
                    if not row_values[3] and len(existing_values) > 3 and existing_values[3]:
                        row_values[3] = existing_values[3]
                    if (row_values[5] is None or row_values[5] == 0) and len(existing_values) > 5 and existing_values[5]:
                        try:
                            row_values[5] = float(existing_values[5])
                        except (TypeError, ValueError):
                            row_values[5] = existing_values[5]
                    if not row_values[6] and len(existing_values) > 6 and existing_values[6]:
                        row_values[6] = existing_values[6]
                updated_entries[match_idx] = {
                    "values": row_values,
                    "link": filepath,
                }
                matched_rows.add(match_idx)
                row_count_updated += 1
            else:
                new_entries.append({
                    "values": row_values,
                    "link": filepath,
                })
                row_count_added += 1

        final_rows = []
        if updated_entries:
            for idx in sorted(updated_entries.keys()):
                final_rows.append(updated_entries[idx])

        if new_entries:
            new_entries.sort(key=lambda item: (item["values"][0] or "", item["values"][1] or ""))
            final_rows.extend(new_entries)

        # Очистка старых строк
        existing_rows_count = ws.max_row - 1
        if existing_rows_count > 0:
            ws.delete_rows(2, existing_rows_count)

        for entry in final_rows:
            ws.append(entry["values"])
            link_cell = ws.cell(row=ws.max_row, column=9)
            link_cell.hyperlink = entry["link"]
            link_cell.font = link_font

        try:
            wb.save(self.xlsx_path)
            self.log(f"Каталог сохранен: {self.xlsx_path}", "SUCCESS")
        except PermissionError:
            self.log("Не удалось сохранить каталог. Файл открыт в другой программе.", "ERROR")
            return

        removed_rows = len(existing_rows) - len(matched_rows) if existing_rows else 0
        if row_count_updated:
            self.log(f"Обновлено записей: {row_count_updated}", "INFO")
        if row_count_added:
            self.log(f"Добавлено записей: {row_count_added}", "INFO")
        if removed_rows:
            self.log(f"Удалено записей: {removed_rows}", "INFO")
