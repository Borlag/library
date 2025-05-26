# =============================  sync_specs.py  =============================
"""
SpecSync v3
-----------

• Работает с PDF и/или растровыми сканами (TIF/PNG/JPG); GUI-чек-бокс «PDF only».
• Категоризация: keyword-rules ▸ OCR (при необходимости) ▸ LLM-fallback.
• Теги: топ-5 технических слов (A-Z0-9-, длина ≥ 3).
• INDEX.xlsx: столбец «Specification Name» (без расширения).
• Поддержка слабого ноутбука: LLM запускается через llama-cpp-python с
  мини-моделью (Q4_K) и двумя потоками CPU; указать путь в переменной `LLAMA_MODEL`.

Пакеты: PySide6, openpyxl, pyyaml, python-dotenv, pdfminer.six, pillow,
pytesseract, pdf2image, llama-cpp-python (по желанию), pywin32 (для .lnk).
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import logging
import os
import re
import shutil
import sqlite3
import sys
import threading
import time
from contextlib import contextmanager
from dataclasses import dataclass
from email.message import EmailMessage
from io import BytesIO
from pathlib import Path
from typing import Callable, Iterator, List, Optional

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import yaml

# GUI
from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtWidgets import (
    QApplication, QFileDialog, QGridLayout, QHBoxLayout, QLabel,
    QLineEdit, QMainWindow, QMenu, QMenuBar, QMessageBox, QPushButton,
    QProgressBar, QRadioButton, QStatusBar, QTextEdit, QWidget, QCheckBox
)

# OCR / PDF helpers
try:
    import pdfminer.high_level as pdfminer
except ImportError:
    pdfminer = None
try:
    from PIL import Image
    import pytesseract
except ImportError:
    pytesseract = None

# Windows shortcuts
if os.name == "nt":
    import pythoncom  # noqa: F401
    import win32com.client  # type: ignore

# ---------------------------  CONSTANTS  ---------------------------------- #
PDF_ONLY = {".pdf"}
ALL_EXTS = {".pdf", ".tif", ".tiff", ".png", ".jpg", ".jpeg"}

REV_RE = re.compile(r"(?:[_\- ]REV[_\-\.\s]?|[_\- ]R)([A-Z])", re.IGNORECASE)

INDEX_COLUMNS = [
    "Abbrev",
    "Specification Name",   # ← без расширения
    "Revision",
    "Category",
    "Tags",
    "Size_MB",
    "Modified",
    "SHA256",
    "Link",
]

DB_SCHEMA = """
CREATE TABLE IF NOT EXISTS files (
    sha256 TEXT PRIMARY KEY,
    abbrev TEXT,
    path TEXT,
    revision TEXT,
    category TEXT,
    tags TEXT,
    size_mb REAL,
    modified TEXT
);
"""

# ---------------------------  DATA CLASS  --------------------------------- #
@dataclass
class FileMeta:
    abbrev: str
    spec_name: str
    revision: str
    category: str
    tags: str
    size_mb: float
    modified: str
    sha256: str
    link: str

    def as_row(self) -> list:
        return [
            self.abbrev,
            self.spec_name,
            self.revision,
            self.category,
            self.tags,
            f"{self.size_mb:.2f}",
            self.modified,
            self.sha256,
            self.link,
        ]

# -------------------------  UTILITY FUNCS  -------------------------------- #
def sha256sum(path: Path, buf: int = 1 << 18) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        while chunk := f.read(buf):
            h.update(chunk)
    return h.hexdigest()


def build_abbrev_regex(words: set[str]) -> re.Pattern:
    esc = sorted((re.escape(w) for w in words), key=len, reverse=True)
    return re.compile("|".join(esc))


def extract_abbrev(name: str, pattern: re.Pattern) -> Optional[str]:
    m = pattern.search(name)
    return m.group(0) if m else None


def extract_revision(name: str) -> str:
    m = REV_RE.search(name)
    return m.group(1).upper() if m else ""


def create_shortcut(src: Path, dst: Path) -> None:
    shell = win32com.client.Dispatch("WScript.Shell")
    sc = shell.CreateShortcut(str(dst))
    sc.TargetPath = str(src)
    sc.WorkingDirectory = str(src.parent)
    sc.IconLocation = str(src)
    sc.save()


def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


@contextmanager
def sqlite_conn(db: Path) -> Iterator[sqlite3.Connection]:
    conn = sqlite3.connect(db)
    conn.execute("PRAGMA journal_mode=WAL")
    try:
        yield conn
    finally:
        conn.commit()
        conn.close()


def send_err_mail(subj: str, body: str) -> None:
    load_dotenv()
    host = os.getenv("SMTP_HOST")
    if not host:
        return
    msg = EmailMessage()
    msg["Subject"] = subj
    msg["From"] = os.getenv("SMTP_USER")
    msg["To"] = os.getenv("SMTP_TO")
    msg.set_content(body)
    import smtplib, ssl
    with smtplib.SMTP_SSL(host, context=ssl.create_default_context()) as s:
        s.login(os.getenv("SMTP_USER"), os.getenv("SMTP_PASS"))
        s.send_message(msg)

# ----------  OCR / TEXT EXTRACTION  -------------------------------------- #
def pdf_text(path: Path) -> str:
    if pdfminer:
        try:
            return pdfminer.extract_text(str(path)) or ""
        except Exception:
            pass
    return ""


def image_ocr(img_bytes: bytes) -> str:
    if pytesseract:
        img = Image.open(BytesIO(img_bytes))
        return pytesseract.image_to_string(img, lang="eng", config="--psm 6")
    return ""


def pdf_ocr(path: Path) -> str:
    if not pytesseract:
        return ""
    try:
        import pdf2image
    except ImportError:
        return ""
    pages = pdf2image.convert_from_path(str(path), dpi=300)
    return "\n".join(pytesseract.image_to_string(p, lang="eng", config="--psm 6")
                     for p in pages)


def extract_text(path: Path) -> str:
    ext = path.suffix.lower()
    if ext == ".pdf":
        txt = pdf_text(path)
        if len(txt) < 30:
            txt = pdf_ocr(path)
        return txt
    if ext in {".tif", ".tiff", ".png", ".jpg", ".jpeg"}:
        try:
            return image_ocr(path.read_bytes())
        except Exception:
            return ""
    return ""


def top_keywords(text: str, k: int = 5) -> List[str]:
    words = re.findall(r"[A-Z][A-Z0-9\-]{2,}", text.upper())
    freq = {}
    for w in words:
        freq[w] = freq.get(w, 0) + 1
    return [w for w, _ in sorted(freq.items(), key=lambda t: (-t[1], t[0]))[:k]]

# ---------------------------  CORE ENGINE  -------------------------------- #
class SpecSync:
    def __init__(
        self,
        source_root: Path,
        target_root: Path,
        abbreviations: Path,
        rules_yaml: Path,
        mode: str = "copy",
        dry_run: bool = False,
        use_llm: bool = False,
        pdf_only: bool = True,
        prog_cb: Callable[[str], None] | None = None,
    ):
        self.src = source_root
        self.dst = target_root
        self.mode = mode
        self.dry = dry_run
        self.use_llm = use_llm
        self.scan_exts = PDF_ONLY if pdf_only else ALL_EXTS
        self.say = prog_cb or (lambda _: None)

        self.whitelist = {ln.strip() for ln in abbreviations.read_text(encoding="utf-8").splitlines() if ln.strip()}
        self.abbrev_pat = build_abbrev_regex(self.whitelist)
        self.rules = yaml.safe_load(rules_yaml.read_text(encoding="utf-8")) or {}

        self.db = self.dst / "inventory.db"
        ensure_dir(self.dst)
        with sqlite_conn(self.db) as c:
            c.executescript(DB_SCHEMA)

        self.s, self.u, self.d, self.e = 0, 0, 0, 0  # counters

    # -----------------------------  public  ------------------------------ #
    def run(self) -> None:
        try:
            self._sync()
        except Exception as exc:
            logging.exception("fatal")
            send_err_mail("SpecSync FAILED", str(exc))
            raise

    # ----------------------------  internal  ----------------------------- #
    def _iter_files(self) -> Iterator[Path]:
        for p in self.src.rglob("*"):
            if p.is_file() and p.suffix.lower() in self.scan_exts:
                self.s += 1
                if extract_abbrev(p.name, self.abbrev_pat):
                    yield p

    def _sync(self) -> None:
        files = list(self._iter_files())
        total = len(files)
        wb, ws = self._prepare_index()
        for i, path in enumerate(files, 1):
            self.say(f"{i}/{total}: {path.name}")
            try:
                self._handle(path, ws)
            except Exception:
                self.e += 1
                logging.exception(path)
        wb.save(self.dst / "INDEX.xlsx")
        self._autosize()
        self.say(f"OK:{self.u}  DUP:{self.d}  ERR:{self.e}")

    # ----------  index helpers  ---------- #
    def _prepare_index(self):
        ix = self.dst / "INDEX.xlsx"
        if ix.exists():
            wb = load_workbook(ix)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(INDEX_COLUMNS)
            ws.auto_filter.ref = f"A1:{get_column_letter(len(INDEX_COLUMNS))}1"
        return wb, ws

    def _autosize(self):
        ix = self.dst / "INDEX.xlsx"
        wb = load_workbook(ix)
        ws = wb.active
        for i, col in enumerate(ws.columns, 1):
            width = min(max(len(str(c.value or "")) for c in col) + 2, 80)
            ws.column_dimensions[get_column_letter(i)].width = width
        wb.save(ix)

    # ----------  per-file  -------------- #
    def _handle(self, src: Path, ws):
        abbrev = extract_abbrev(src.name, self.abbrev_pat)
        sha = sha256sum(src)
        with sqlite_conn(self.db) as c:
            if c.execute("SELECT 1 FROM files WHERE sha256=?", (sha,)).fetchone():
                self.d += 1
                if self.mode == "move_link" and not self.dry:
                    target = Path(c.execute("SELECT path FROM files WHERE sha256=?", (sha,)).fetchone()[0])
                    self._shortcut(src, target)
                return
        # unique
        self.u += 1
        rev = extract_revision(src.stem)
        text = extract_text(src)
        cat, tags = self._categorise(src.name, text)
        dest = self._place(src, abbrev)
        meta = FileMeta(
            abbrev, dest.stem, rev, cat, ",".join(tags),
            src.stat().st_size / 1048576,
            dt.datetime.fromtimestamp(src.stat().st_mtime, dt.timezone.utc).isoformat(),
            sha,
            f'=HYPERLINK("{dest.as_posix()}", "{dest.name}")',
        )
        ws.append(meta.as_row())
        with sqlite_conn(self.db) as c:
            c.execute(
                "INSERT INTO files VALUES (?,?,?,?,?,?,?,?)",
                (sha, abbrev, str(dest), rev, cat, meta.tags, meta.size_mb, meta.modified),
            )

    def _place(self, src: Path, abbrev: str) -> Path:
        dir_ = self.dst / abbrev
        ensure_dir(dir_)
        dest = dir_ / src.name
        if dest.exists():
            b, e = os.path.splitext(src.name)
            n = 1
            while (dir_ / f"{b}__{n}{e}").exists():
                n += 1
            dest = dir_ / f"{b}__{n}{e}"
        if not self.dry:
            if self.mode == "copy":
                shutil.copy2(src, dest)
            else:
                shutil.move(src, dest)
                self._shortcut(src, dest)
        return dest

    def _shortcut(self, orig: Path, tgt: Path) -> None:
        if os.name != "nt":
            return
        lnk = orig.with_suffix(".lnk")
        if not lnk.exists():
            create_shortcut(tgt, lnk)

    # ----------  categorisation  ---------- #
    def _categorise(self, fn: str, text: str) -> tuple[str, List[str]]:
        fn_low = fn.lower()
        for cat, kw in self.rules.items():
            if any(k.lower() in fn_low for k in kw):
                return cat, kw
        if text:
            kws = top_keywords(text)
            for cat, kw in self.rules.items():
                if any(k in text.upper() for k in kw):
                    return cat, kws
            if kws:
                return "Text-Derived", kws
        if self.use_llm:
            return self._llm(fn, text)
        return "Uncategorised", []

    def _llm(self, fn: str, txt: str) -> tuple[str, List[str]]:
        try:
            from llama_cpp import Llama
            model = Path(os.getenv("LLAMA_MODEL", "")).expanduser()
            if not model.exists():
                return "LLM-Unk", []
            ll = Llama(model_path=str(model), n_ctx=1024, n_threads=2)
            prompt = (
                "You are a file-classifier. Return one short category name and up to "
                "five comma-separated tags.\n"
                f"File name: {fn}\n"
                f"Content sample:\n{txt[:1500]}\n\nAnswer:"
            )
            out = ll(prompt, max_tokens=64, stop=["</s>"])["choices"][0]["text"].strip()
            if "\n" in out:
                cat, tag_line = out.split("\n", 1)
            else:
                cat, tag_line = out, ""
            tags = [t.strip("# ,") for t in re.split(r"[;,]", tag_line) if t.strip()]
            return cat or "LLM", tags
        except Exception as e:
            logging.warning("LLM failed: %s", e)
            return "LLM-Unk", []

# ---------------------------  GUI  ---------------------------------------- #
class Worker(QThread):
    prog = Signal(str)
    done = Signal()

    def __init__(self, eng: SpecSync):
        super().__init__()
        self.eng = eng

    def run(self):
        self.eng.say = self.prog.emit
        self.eng.run()
        self.done.emit()


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("SpecSync v3")
        self.setMinimumSize(960, 640)
        self._ui()
        self.worker: Optional[Worker] = None
        logging.basicConfig(filename="sync.log",
                            level=logging.INFO,
                            format="%(asctime)s %(levelname)s %(message)s")

    # UI
    def _ui(self):
        w = QWidget()
        self.setCentralWidget(w)
        g = QGridLayout(w)

        self.srcE = QLineEdit(); self.dstE = QLineEdit()
        self.abbrE = QLineEdit(); self.rulesE = QLineEdit("rules.yaml")
        for r, (lab, edit, kind) in enumerate(
            [("Source Root", self.srcE, "dir"),
             ("Target Root", self.dstE, "dir"),
             ("abbreviations.txt", self.abbrE, "file"),
             ("rules.yaml", self.rulesE, "file")]
        ):
            g.addWidget(QLabel(lab), r, 0)
            g.addWidget(edit, r, 1)
            b = QPushButton("…"); g.addWidget(b, r, 2)
            b.clicked.connect(lambda _, e=edit, k=kind: self._pick(e, k))

        self.copyR, self.moveR = QRadioButton("Copy"), QRadioButton("Move & Link")
        self.copyR.setChecked(True)
        box = QHBoxLayout(); box.addWidget(self.copyR); box.addWidget(self.moveR)
        g.addLayout(box, 4, 0, 1, 3)

        self.llmC = QCheckBox("Use LLM fallback")
        self.dryC = QCheckBox("Dry-Run")
        self.pdfC = QCheckBox("PDF only"); self.pdfC.setChecked(True)
        box2 = QHBoxLayout()
        box2.addWidget(self.llmC); box2.addWidget(self.dryC); box2.addWidget(self.pdfC)
        g.addLayout(box2, 5, 0, 1, 3)

        self.pbar = QProgressBar(); self.pbar.setRange(0, 0); self.pbar.hide()
        self.logE = QTextEdit(); self.logE.setReadOnly(True)
        g.addWidget(self.pbar, 6, 0, 1, 3)
        g.addWidget(self.logE, 7, 0, 1, 3)

        self.startB = QPushButton("Start"); self.cancelB = QPushButton("Cancel")
        self.cancelB.setEnabled(False)
        box3 = QHBoxLayout(); box3.addWidget(self.startB); box3.addWidget(self.cancelB)
        g.addLayout(box3, 8, 0, 1, 3)
        self.startB.clicked.connect(self._start); self.cancelB.clicked.connect(self._cancel)

        self.status = QStatusBar(); self.setStatusBar(self.status)

        # menu Safety
        mb = QMenuBar(); tools = QMenu("Tools", self); safety = tools.addMenu("Safety")
        safety.addAction("Create Dump", self._dump)
        safety.addAction("Restore from Dump", self._restore)
        mb.addMenu(tools); self.setMenuBar(mb)

    def _pick(self, edit: QLineEdit, kind: str):
        if kind == "dir":
            d = QFileDialog.getExistingDirectory(self, "Choose")
            if d: edit.setText(d)
        else:
            f, _ = QFileDialog.getOpenFileName(self, "Choose file")
            if f: edit.setText(f)

    # Sync control
    def _start(self):
        if self.worker and self.worker.isRunning():
            return
        src, dst = Path(self.srcE.text()), Path(self.dstE.text())
        abbr, rules = Path(self.abbrE.text()), Path(self.rulesE.text())
        if not (src.is_dir() and dst and abbr.exists() and rules.exists()):
            QMessageBox.critical(self, "", "Select valid paths")
            return
        eng = SpecSync(
            src, dst, abbr, rules,
            mode="copy" if self.copyR.isChecked() else "move_link",
            dry_run=self.dryC.isChecked(), use_llm=self.llmC.isChecked(),
            pdf_only=self.pdfC.isChecked(),
        )
        self.worker = Worker(eng)
        self.worker.prog.connect(self._log)
        self.worker.done.connect(self._done)
        self.worker.start()
        self._busy(True)
        self.logE.clear(); self._log("== START ==")

    def _cancel(self):
        if self.worker and self.worker.isRunning():
            self.worker.terminate()
            self.worker.wait()
        self._done()

    def _done(self):
        self._busy(False); self._log("== FINISHED ==")

    def _busy(self, flag: bool):
        self.startB.setEnabled(not flag); self.cancelB.setEnabled(flag)
        self.pbar.setVisible(flag)

    def _log(self, s: str):
        ts = time.strftime("%H:%M:%S"); self.logE.append(f"[{ts}] {s}")
        self.status.showMessage(s); logging.info(s)

    # Safety
    def _dump(self):
        dst = Path(self.dstE.text()); db = dst / "inventory.db"
        if not db.exists():
            return
        out = dst / f"dump_{dt.datetime.now():%Y%m%d}.json"
        with sqlite_conn(db) as c, out.open("w", encoding="utf-8") as f:
            cols = [d[0] for d in c.execute("PRAGMA table_info(files)")]
            json.dump([dict(zip(cols, r)) for r in c.execute("SELECT * FROM files")], f, indent=2)
        QMessageBox.information(self, "", f"Dump → {out}")

    def _restore(self):
        dump, _ = QFileDialog.getOpenFileName(self, "dump_*.json", filter="JSON (*.json)")
        if not dump:
            return
        dst = Path(self.dstE.text()); db = dst / "inventory.db"
        with sqlite_conn(db) as c, open(dump, encoding="utf-8") as f:
            data = json.load(f); c.execute("DELETE FROM files")
            for r in data:
                c.execute("INSERT INTO files VALUES (?,?,?,?,?,?,?,?)",
                          (r['sha256'], r['abbrev'], r['path'], r['revision'],
                           r['category'], r['tags'], r['size_mb'], r['modified']))
        QMessageBox.information(self, "", "Restored")

# ---------------------------  CLI  ---------------------------------------- #
def cli():
    p = argparse.ArgumentParser("Spec repository sync (headless)")
    p.add_argument("source", type=Path); p.add_argument("target", type=Path)
    p.add_argument("--abbr", type=Path, default="abbreviations.txt")
    p.add_argument("--rules", type=Path, default="rules.yaml")
    p.add_argument("--mode", choices=("copy", "move_link"), default="copy")
    p.add_argument("--dry-run", action="store_true"); p.add_argument("--use-llm", action="store_true")
    p.add_argument("--all-ext", action="store_true", help="scan all supported formats")
    a = p.parse_args()
    logging.basicConfig(filename="sync.log", level=logging.INFO,
                        format="%(asctime)s %(levelname)s %(message)s")
    SpecSync(a.source, a.target, a.abbr, a.rules,
             mode=a.mode, dry_run=a.dry_run, use_llm=a.use_llm,
             pdf_only=not a.all_ext).run()

# ---------------------------  MAIN  --------------------------------------- #
if __name__ == "__main__":
    if len(sys.argv) > 1 and Path(sys.argv[1]).exists():
        cli()
    else:
        app = QApplication(sys.argv)
        mw = MainWindow(); mw.show()
        sys.exit(app.exec())
